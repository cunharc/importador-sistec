from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import csv
import os

def listar_abas(caminho_arquivo: str) -> list:
    wb = load_workbook(caminho_arquivo, read_only=True, data_only=True)
    abas = wb.sheetnames
    wb.close()
    return abas

def ler_planilha(caminho_arquivo: str, aba: str, coluna_conta: str, 
                 coluna_descricao: str, linha_inicial: int) -> list:
    """
    Retorna lista de dicts: [{'conta': '1.1.1.01', 'descricao': 'CAIXA'}, ...]
    """
    wb = load_workbook(caminho_arquivo, read_only=True, data_only=True)
    ws = wb[aba]
    
    col_conta_idx = column_index_from_string(coluna_conta)
    col_desc_idx = column_index_from_string(coluna_descricao)
    
    registros = []
    for row in ws.iter_rows(min_row=linha_inicial, values_only=True):
        if len(row) < max(col_conta_idx, col_desc_idx):
            continue
            
        conta = row[col_conta_idx - 1]
        descricao = row[col_desc_idx - 1]
        
        if conta is None or str(conta).strip() == '':
            continue
            
        registros.append({
            'conta': str(conta).strip(),
            'descricao': str(descricao).strip() if descricao else ''
        })
    
    wb.close()
    return registros

def obter_abas_planilha(caminho_arquivo: str) -> list:
    """Lê as abas de forma segura suportando XLSX e CSV."""
    if str(caminho_arquivo).lower().endswith('.csv'):
        return ['CSV (Aba Única)']
    try:
        wb = load_workbook(caminho_arquivo, read_only=True, data_only=True)
        abas = wb.sheetnames
        wb.close()
        return abas
    except Exception:
        return []

def _detect_csv_dialect(caminho_arquivo: str) -> tuple:
    """Auto-detect CSV delimiter and encoding."""
    delimiters = [',', ';', '\t']

    def _count_unquoted(text: str, delim: str) -> int:
        count = 0
        in_quotes = False
        for ch in text:
            if ch == '"':
                in_quotes = not in_quotes
            elif ch == delim and not in_quotes:
                count += 1
        return count

    def _detect_encoding(path: str) -> str:
        for enc in ['cp1252', 'utf-8-sig', 'latin-1']:
            try:
                with open(path, 'r', encoding=enc) as f:
                    f.read(1024)
                return enc
            except (UnicodeDecodeError, Exception):
                continue
        return 'cp1252'

    encoding = _detect_encoding(caminho_arquivo)
    try:
        with open(caminho_arquivo, 'r', encoding=encoding) as f:
            first_line = f.readline()
    except Exception:
        return ',', encoding

    best = max(delimiters, key=lambda d: _count_unquoted(first_line, d))
    if _count_unquoted(first_line, best) == 0:
        best = ','
    return best, encoding


def _split_codigo_descricao(reg: dict, mapa_colunas: dict):
    """Se codigo e descricao apontam pra mesma coluna, separa por linha, classificando digitos como codigo e resto como descricao."""
    cod_letra = mapa_colunas.get('codigo', '').strip().upper()
    desc_letra = mapa_colunas.get('descricao', '').strip().upper()
    if not cod_letra or not desc_letra or cod_letra != desc_letra:
        return
    raw = reg.get('codigo', reg.get('descricao', ''))
    parts = [p.strip() for p in raw.replace('\r\n', '\n').split('\n') if p.strip()]
    if len(parts) <= 1:
        return
    cod_parts, desc_parts = [], []
    for p in parts:
        if p.isdigit() or (p.startswith('-') and p[1:].isdigit()):
            cod_parts.append(p)
        else:
            desc_parts.append(p)
    reg['codigo'] = ' '.join(cod_parts) if cod_parts else parts[0]
    reg['descricao'] = ' '.join(desc_parts) if desc_parts else parts[-1]


def ler_planilha_produtos(caminho_arquivo: str, aba: str, mapa_colunas: dict, linha_inicial: int) -> list:
    """Lê CSVs ou Excel com base num dicionário de letras de colunas: {'descricao': 'C', 'grupo': 'D'}"""
    registros = []
    
    if caminho_arquivo.lower().endswith('.csv'):
        delimiter, encoding = _detect_csv_dialect(caminho_arquivo)
        with open(caminho_arquivo, 'r', encoding=encoding, errors='ignore') as f:
            reader = csv.reader(f, delimiter=delimiter)
            for i, row in enumerate(reader, start=1):
                if i < linha_inicial: continue
                if not row: continue
                reg = {}
                for chave, letra in mapa_colunas.items():
                    if not letra:
                        reg[chave] = ''
                        continue
                    # Converte Letra da Coluna para Índice Numérico
                    idx = 0
                    for char in letra.strip().upper():
                        idx = idx * 26 + (ord(char) - ord('A')) + 1
                    idx -= 1
                    reg[chave] = str(row[idx]).strip() if idx < len(row) else ''
                _split_codigo_descricao(reg, mapa_colunas)
                if any(reg.values()): registros.append(reg)
    else:
        wb = load_workbook(caminho_arquivo, read_only=True, data_only=True)
        ws = wb[aba] if aba in wb.sheetnames else wb.active
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i < linha_inicial: continue
            reg = {}
            for chave, letra in mapa_colunas.items():
                if not letra:
                    reg[chave] = ''
                    continue
                idx = column_index_from_string(letra.strip().upper()) - 1
                val = row[idx] if idx < len(row) else ''
                reg[chave] = str(val).strip() if val is not None else ''
            _split_codigo_descricao(reg, mapa_colunas)
            if any(reg.values()): registros.append(reg)
        wb.close()
    return registros
