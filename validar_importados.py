# -*- coding: utf-8 -*-
"""
Valida o arquivo IMPORTADOS.xlsx exportado do ERP.

Uso:
    python validar_importados.py "C:\\Users\\SEU_USUARIO\\Desktop\\IMPORTADOS.xlsx"

Mostra, para cada aba:
  - o cabecalho (letra da coluna -> titulo)
  - a quantidade de linhas
  - a soma de cada coluna que parece numerica (valores)

Assim conseguimos ver qual coluna e "Valor" e qual e "Recebido" e conferir os totais.
"""
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def _to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    s = s.replace('R$', '').replace(' ', '')
    # formato BR: 1.234,56 -> 1234.56
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def main(caminho):
    wb = load_workbook(caminho, read_only=True, data_only=True)
    for aba in wb.sheetnames:
        ws = wb[aba]
        print("=" * 70)
        print(f"ABA: {aba}")
        print("=" * 70)

        linhas = list(ws.iter_rows(values_only=True))
        if not linhas:
            print("  (vazia)")
            continue

        header = linhas[0]
        dados = linhas[1:]
        print(f"Linhas de dados: {len(dados)}")
        print("\nColunas:")
        somas = {}
        contagem_num = {}
        for i, titulo in enumerate(header):
            letra = get_column_letter(i + 1)
            print(f"  {letra}: {titulo}")
            somas[i] = 0.0
            contagem_num_i = 0
            for row in dados:
                if i < len(row):
                    f = _to_float(row[i])
                    if f is not None:
                        somas[i] += f
                        contagem_num_i += 1
            contagem_num[i] = contagem_num_i

        print("\nSomas das colunas numericas (>=1 numero):")
        for i, titulo in enumerate(header):
            if contagem_num.get(i, 0) > 0:
                letra = get_column_letter(i + 1)
                print(f"  {letra} ({titulo}): soma = {somas[i]:,.2f}  ({contagem_num[i]} valores)")
        print()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python validar_importados.py \"CAMINHO\\IMPORTADOS.xlsx\"")
        sys.exit(1)
    main(sys.argv[1])
