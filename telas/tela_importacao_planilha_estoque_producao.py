import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import configparser
import threading
import datetime
import csv
import re
import os
import sys

from utils import tema
from utils.excel_reader import obter_abas_planilha, ler_planilha_produtos
from utils.firebird_service import FirebirdService


# Mapeamento de colunas da planilha -> campos da pesagem (etiqueta)
CAMPOS_MAPEAMENTO = [
    ("Etiqueta *", "etiqueta", True),
    ("Produto *", "produto", True),
    ("Peso Líquido *", "peso", True),
    ("Descrição", "descricao", False),
    ("Peso Bruto", "peso_bruto", False),
    ("Qtde Peças", "qtde", False),
    ("Tara Embalagem", "tara_emb", False),
    ("Tara Caixa", "tara_caixa", False),
    ("Data/Hora Pesagem", "data_hora", False),
    ("Validade", "validade", False),
    ("Produção", "producao", False),
]

# Valores fixos validados contra o banco (FRIGODIL.FDB)
COD_TIPO_DESOSSA = 'D'
CDA_QTDE_PADRAO = 999999
USUARIO_PESAGEM_PADRAO = 'SISTEC_MASTER'
OBS_PADRAO = 'inventario importado'
# COD_STATUS e VARCHAR(1) no banco: 'PO' e rejeitado ("expected 1, found 2").
# O ERP usa A=Aberta, P=Produzindo, F=Finalizada (VIEW_DASH_ORDEM_PRODUCAO).
COD_STATUS_PADRAO = 'P'
LOTE_COMMIT = 300

# Como a coluna "Produto" da planilha e casada com o cadastro do ERP.
# No modo automatico a busca segue esta ordem ate achar.
BUSCA_PRODUTO = [
    ("Automático (código, importação, auxiliar)", "auto"),
    ("Código do produto (ERP)", "codigo"),
    ("Código de importação", "importacao"),
    ("Código auxiliar", "auxiliar"),
]
CAMPO_ERP_PRODUTO = {
    'codigo': 'PRODUTO_CODIGO',
    'importacao': 'PRODUTO_COD_IMPORTACAO',
    'auxiliar': 'PRODUTO_COD_AUXILIAR',
}
ROTULO_ORIGEM = {'codigo': 'CÓDIGO', 'importacao': 'IMPORTAÇÃO', 'auxiliar': 'AUXILIAR'}


class TelaImportacaoPlanilhaEstoqueProducao(ttk.Frame):
    """Importa estoque de Produto Acabado (PA) por etiquetas para o ERP.

    Gera a cadeia CONFRI_ORDEM_DESOSSA -> CONFRI_ORDEM_DESOSSA_PA ->
    CONFRI_ORDEM_DESOSSA_PA_PESAGEM: uma ordem de inventario, um item PA por
    produto distinto e uma pesagem por etiqueta da planilha.
    """

    def __init__(self, parent, callback_voltar=None):
        super().__init__(parent, padding="10")
        self.parent = parent
        self.callback_voltar = callback_voltar
        self.pack(fill=tk.BOTH, expand=True)

        self.registros_lidos = []
        self.caminho_arquivo = ""
        self.dados_grid = {}
        self._sort_directions = {}
        self._clientes = []

        self.config = configparser.ConfigParser()
        self.config.read('config.ini', encoding='utf-8')
        # o fbclient do INI vem como nome simples (fbclient_5.dll); sem resolver para
        # caminho absoluto o fdb nao acha a DLL e a conexao falha
        fbcli = self.config.get('FIREBIRD', 'fbclient', fallback='').strip()
        self.config_db = {
            'host': self.config.get('FIREBIRD', 'servidor', fallback='127.0.0.1'),
            'port': self.config.get('FIREBIRD', 'porta', fallback='3050'),
            'database': self.config.get('FIREBIRD', 'caminho_banco', fallback=''),
            'user': self.config.get('FIREBIRD', 'usuario', fallback='SYSDBA'),
            'password': self.config.get('FIREBIRD', 'senha', fallback='masterkey'),
            'fbclient': self._resource_path(fbcli) if fbcli else ''
        }

        self._criar_widgets()
        self._carregar_config_mapeamento()
        self._carregar_clientes_bg()

    def _resource_path(self, relative_path):
        """Resolve arquivo do projeto, funcionando no dev e no PyInstaller."""
        if os.path.isabs(relative_path):
            return relative_path
        try:
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)

    # ------------------------------------------------------------------ UI
    def _criar_widgets(self):
        tema.montar_header(
            self, "Importar Estoque de Produção (Excel)",
            "Estoque de Produto Acabado por etiquetas: gera Ordem de Desossa, itens PA e as pesagens"
        ).pack(fill=tk.X)

        corpo = tk.Frame(self, bg=tema.BG_BASE)
        corpo.pack(fill=tk.BOTH, expand=True)

        sidebar = tema.montar_sidebar(corpo)

        rodape_sb = tk.Frame(sidebar, bg=tema.SIDEBAR_BG)
        rodape_sb.pack(side=tk.BOTTOM, fill=tk.X, pady=(0, 8))
        self.btn_voltar = tema.botao_sidebar(rodape_sb, "⎋   Voltar", self._fechar_tela)
        self.btn_voltar.pack(fill=tk.X)

        tema.titulo_sidebar(sidebar, "AÇÕES").pack(fill=tk.X, pady=(16, 4))

        self.btn_analisar = tema.botao_sidebar(sidebar, "🔍   Carregar e Analisar Planilha", self._iniciar_analise)
        self.btn_analisar.pack(fill=tk.X)

        self.btn_importar = tema.botao_sidebar(sidebar, "🚀   Importar Estoque no ERP", self._iniciar_importacao, cor_fg="#7EE0A0")
        self.btn_importar.config(state=tk.DISABLED)
        self.btn_importar.pack(fill=tk.X)

        self.btn_exportar = tema.botao_sidebar(sidebar, "📋   Exportar Análise (Excel/CSV)", self._exportar_analise, cor_fg="#8FD8FF")
        self.btn_exportar.config(state=tk.DISABLED)
        self.btn_exportar.pack(fill=tk.X)

        self.btn_exportar_faltantes = tema.botao_sidebar(
            sidebar, "📦   Exportar Produtos a Cadastrar", self._exportar_faltantes, cor_fg="#FFC48F")
        self.btn_exportar_faltantes.config(state=tk.DISABLED)
        self.btn_exportar_faltantes.pack(fill=tk.X)

        content = tk.Frame(corpo, bg=tema.BG_BASE)
        content.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=16, pady=12)

        # --- CARDS DE RESUMO ---
        frame_cards = ttk.Frame(content)
        frame_cards.pack(fill=tk.X, pady=(8, 2), padx=5)

        self.card_etiquetas = self._criar_card(frame_cards, "Etiquetas a Importar", "0", "#14146E")
        self.card_etiquetas.pack(side=tk.LEFT, padx=5)

        self.card_peso = self._criar_card(frame_cards, "Peso Total (kg)", "0,000", "#22C55E")
        self.card_peso.pack(side=tk.LEFT, padx=5)

        self.card_produtos = self._criar_card(frame_cards, "Produtos (itens PA)", "0", "#E67E22")
        self.card_produtos.pack(side=tk.LEFT, padx=5)

        self.card_erros = self._criar_card(frame_cards, "Com Problema", "0", "#C8001E")
        self.card_erros.pack(side=tk.LEFT, padx=5)

        # --- ARQUIVO ---
        file_row = ttk.Frame(content)
        file_row.pack(fill=tk.X, pady=2)

        tk.Label(file_row, text="Arquivo:", font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT, padx=(5, 2))
        self.ent_arquivo = ttk.Entry(file_row, font=("Segoe UI", 9))
        self.ent_arquivo.pack(side=tk.LEFT, padx=2, fill=tk.X, expand=True)
        ttk.Button(file_row, text="📁 Selecionar", command=self._selecionar_arquivo).pack(side=tk.LEFT, padx=2)

        tk.Label(file_row, text="Aba:", font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT, padx=(10, 2))
        self.cb_abas = ttk.Combobox(file_row, width=16, state="readonly", font=("Segoe UI", 9))
        self.cb_abas.pack(side=tk.LEFT, padx=2)

        tk.Label(file_row, text="Linha Inicial:", font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT, padx=(10, 2))
        self.ent_linha_ini = ttk.Entry(file_row, width=6, font=("Segoe UI", 9))
        self.ent_linha_ini.insert(0, "2")
        self.ent_linha_ini.pack(side=tk.LEFT, padx=2)

        # --- MAPEAMENTO ---
        frame_map = ttk.LabelFrame(content, text="Mapeamento de Colunas (Insira a letra: A, B, C...)", padding="8")
        frame_map.pack(fill=tk.X, pady=4)

        self.entradas_map = {}
        linhas_campos = [CAMPOS_MAPEAMENTO[i:i + 4] for i in range(0, len(CAMPOS_MAPEAMENTO), 4)]
        for i_linha, grupo in enumerate(linhas_campos):
            for i_col, (lbl_texto, chave, obrigatorio) in enumerate(grupo):
                col = i_col * 2
                fg_color = "#C8001E" if obrigatorio else "#1A1A1A"
                tk.Label(frame_map, text=lbl_texto, font=("Segoe UI", 8, "bold"),
                         fg=fg_color).grid(row=i_linha, column=col, padx=(5, 1), pady=2, sticky=tk.E)
                ent = ttk.Entry(frame_map, width=5, font=("Segoe UI", 9))
                ent.grid(row=i_linha, column=col + 1, padx=(0, 5), pady=2, sticky=tk.W)
                self.entradas_map[chave] = ent

        linha_busca = len(linhas_campos)
        tk.Label(frame_map, text="Casar 'Produto' por:", font=("Segoe UI", 8, "bold"),
                 fg="#14146E").grid(row=linha_busca, column=0, padx=(5, 1), pady=(6, 2), sticky=tk.E)
        self.cb_busca_produto = ttk.Combobox(
            frame_map, width=38, state="readonly", font=("Segoe UI", 9),
            values=[lbl for lbl, _ in BUSCA_PRODUTO])
        self.cb_busca_produto.current(0)
        self.cb_busca_produto.grid(row=linha_busca, column=1, columnspan=4,
                                   padx=(0, 5), pady=(6, 2), sticky=tk.W)
        tk.Label(frame_map, text="(aceita o código do ERP, o código de importação ou o auxiliar)",
                 font=("Segoe UI", 8), fg="#555").grid(row=linha_busca, column=5, columnspan=3,
                                                       padx=5, pady=(6, 2), sticky=tk.W)

        # --- PARAMETROS DA ORDEM ---
        frame_ordem = ttk.LabelFrame(content, text="Ordem de Inventário (CONFRI_ORDEM_DESOSSA)", padding="8")
        frame_ordem.pack(fill=tk.X, pady=4)

        tk.Label(frame_ordem, text="Empresa:", font=("Segoe UI", 8, "bold")).grid(row=0, column=0, padx=(5, 1), pady=2, sticky=tk.E)
        self.ent_empresa = ttk.Entry(frame_ordem, width=5, font=("Segoe UI", 9))
        self.ent_empresa.insert(0, self.config.get('IMPORTACAO', 'empresa', fallback='1'))
        self.ent_empresa.grid(row=0, column=1, padx=(0, 5), pady=2, sticky=tk.W)

        tk.Label(frame_ordem, text="Filial:", font=("Segoe UI", 8, "bold")).grid(row=0, column=2, padx=(5, 1), pady=2, sticky=tk.E)
        self.ent_filial = ttk.Entry(frame_ordem, width=5, font=("Segoe UI", 9))
        self.ent_filial.insert(0, self.config.get('IMPORTACAO', 'filial', fallback='1'))
        self.ent_filial.grid(row=0, column=3, padx=(0, 5), pady=2, sticky=tk.W)

        tk.Label(frame_ordem, text="Data da Ordem:", font=("Segoe UI", 8, "bold")).grid(row=0, column=4, padx=(5, 1), pady=2, sticky=tk.E)
        self.ent_data_ordem = ttk.Entry(frame_ordem, width=12, font=("Segoe UI", 9))
        self.ent_data_ordem.insert(0, datetime.date.today().strftime('%d/%m/%Y'))
        self.ent_data_ordem.grid(row=0, column=5, padx=(0, 5), pady=2, sticky=tk.W)

        tk.Label(frame_ordem, text="Status:", font=("Segoe UI", 8, "bold")).grid(row=0, column=6, padx=(5, 1), pady=2, sticky=tk.E)
        self.cb_status = ttk.Combobox(frame_ordem, width=4, state="readonly", font=("Segoe UI", 9),
                                      values=["P", "A", "F"])
        self.cb_status.set(COD_STATUS_PADRAO)
        self.cb_status.grid(row=0, column=7, padx=(0, 5), pady=2, sticky=tk.W)

        tk.Label(frame_ordem, text="Cliente padrão de estoque:", font=("Segoe UI", 8, "bold")).grid(row=1, column=0, columnspan=2, padx=(5, 1), pady=2, sticky=tk.E)
        self.cb_cliente = ttk.Combobox(frame_ordem, width=44, state="readonly", font=("Segoe UI", 9))
        self.cb_cliente.grid(row=1, column=2, columnspan=4, padx=(0, 5), pady=2, sticky=tk.W)

        tk.Label(frame_ordem, text="Observação:", font=("Segoe UI", 8, "bold")).grid(row=1, column=6, padx=(5, 1), pady=2, sticky=tk.E)
        self.ent_obs = ttk.Entry(frame_ordem, width=26, font=("Segoe UI", 9))
        self.ent_obs.insert(0, OBS_PADRAO)
        self.ent_obs.grid(row=1, column=7, columnspan=2, padx=(0, 5), pady=2, sticky=tk.W)

        tk.Label(frame_ordem, text="Usuário da pesagem:", font=("Segoe UI", 8, "bold")).grid(row=2, column=0, columnspan=2, padx=(5, 1), pady=2, sticky=tk.E)
        self.ent_usuario = ttk.Entry(frame_ordem, width=20, font=("Segoe UI", 9))
        self.ent_usuario.insert(0, USUARIO_PESAGEM_PADRAO)
        self.ent_usuario.grid(row=2, column=2, columnspan=2, padx=(0, 5), pady=2, sticky=tk.W)

        self.var_uma_ordem = tk.BooleanVar(value=True)
        ttk.Checkbutton(frame_ordem, text="Uma única ordem para toda a planilha",
                        variable=self.var_uma_ordem).grid(row=2, column=4, columnspan=4, padx=5, pady=2, sticky=tk.W)

        # --- ACOES / STATUS ---
        actions_row = ttk.Frame(content)
        actions_row.pack(fill=tk.X, pady=4)

        ttk.Button(actions_row, text="☑ Marcar Todos", command=self._marcar_todos).pack(side=tk.LEFT, padx=3)
        ttk.Button(actions_row, text="☐ Desmarcar", command=self._desmarcar_todos).pack(side=tk.LEFT, padx=3)

        self.progresso = ttk.Progressbar(actions_row, orient=tk.HORIZONTAL, mode='determinate', length=120)
        self.progresso.pack(side=tk.LEFT, padx=8)

        self.lbl_status = ttk.Label(actions_row, text="Aguardando configuração...", font=("Segoe UI", 9), foreground="#555")
        self.lbl_status.pack(side=tk.LEFT, padx=2)

        filter_row = ttk.Frame(content)
        filter_row.pack(fill=tk.X, pady=(2, 0))

        tk.Label(filter_row, text="Filtrar Status:", font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT, padx=(5, 2))
        self.cb_filtro_status = ttk.Combobox(
            filter_row, state="readonly", width=26, font=("Segoe UI", 9),
            values=["Todos", "OK", "JÁ CADASTRADA", "DUPLICADA",
                    "PRODUTO NÃO ENCONTRADO", "PRODUTO AMBÍGUO", "ERRO"])
        self.cb_filtro_status.current(0)
        self.cb_filtro_status.pack(side=tk.LEFT, padx=2)
        self.cb_filtro_status.bind("<<ComboboxSelected>>", self._filtrar_status)

        # --- GRADE ---
        frame_grade = ttk.Frame(content)
        frame_grade.pack(fill=tk.BOTH, expand=True, pady=4)

        self.colunas = ("SEL", "STATUS", "ETIQUETA", "PROD. PLANILHA", "COD. ERP", "CASOU POR",
                        "DESCRIÇÃO", "PESO", "BRUTO", "QTDE", "PRODUÇÃO", "VENCIMENTO",
                        "DIAS VAL.")
        self._sort_directions = {col: False for col in self.colunas}
        self.tree = ttk.Treeview(frame_grade, columns=self.colunas, show="headings")
        self.tree.bind("<ButtonRelease-1>", self._on_tree_click)
        self.tree.bind("<Double-1>", self._on_tree_double_click)
        self.tree.bind("<Control-c>", self._on_ctrl_c)
        self.tree.bind("<Control-C>", self._on_ctrl_c)

        larguras = [40, 160, 140, 100, 75, 95, 200, 75, 75, 55, 85, 85, 70]
        alinhar_esq = ("ETIQUETA", "DESCRIÇÃO")
        for col, larg in zip(self.colunas, larguras):
            self.tree.heading(col, text=col + " ↕", command=lambda c=col: self._sort_treeview(c))
            self.tree.column(col, width=larg, anchor=tk.W if col in alinhar_esq else tk.CENTER)

        tk.Label(content, text="Dica: dê duplo-clique na ETIQUETA para copiar o código "
                               "(Ctrl+C também copia a linha selecionada).   "
                               "DIAS VAL. = VENCIMENTO − PRODUÇÃO (informativo, não vai para o ERP).",
                 font=("Segoe UI", 8), bg=tema.BG_BASE, fg="#555").pack(anchor=tk.W, padx=5)

        self.tree.tag_configure('OK', background='#EAFAF1', foreground='#16A34A')
        self.tree.tag_configure('AVISO', background='#FFF8E1', foreground='#B45309')
        self.tree.tag_configure('ERRO', background='#FDECEA', foreground='#C8001E')

        scroll_y = ttk.Scrollbar(frame_grade, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=scroll_y.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)

    def _criar_card(self, parent, titulo, valor_inicial, cor_texto):
        """Cria um card de resumo para os totais."""
        card = tk.Frame(parent, bg="#FFFFFF", highlightbackground="#CCCCCC", highlightthickness=1, padx=15, pady=8)
        tk.Label(card, text=titulo, font=("Segoe UI", 9, "bold"), bg="#FFFFFF", fg="#555").pack(anchor=tk.E)
        lbl_valor = tk.Label(card, text=valor_inicial, font=("Segoe UI", 14, "bold"), bg="#FFFFFF", fg=cor_texto)
        lbl_valor.pack(anchor=tk.E)
        card.lbl_valor = lbl_valor
        return card

    # -------------------------------------------------------------- CONFIG
    def _salvar_config_mapeamento(self):
        secao = 'IMPORTACAO_ESTOQUE_PRODUCAO'
        if not self.config.has_section(secao):
            self.config.add_section(secao)
        self.config.set(secao, 'ultimo_arquivo', self.caminho_arquivo)
        self.config.set(secao, 'ultima_aba', self.cb_abas.get())
        self.config.set(secao, 'linha_inicial', self.ent_linha_ini.get())
        for chave, ent in self.entradas_map.items():
            self.config.set(secao, f'map_{chave}', ent.get().strip())
        self.config.set(secao, 'empresa', self.ent_empresa.get().strip())
        self.config.set(secao, 'filial', self.ent_filial.get().strip())
        self.config.set(secao, 'cod_status', self.cb_status.get().strip())
        self.config.set(secao, 'observacao', self.ent_obs.get().strip())
        self.config.set(secao, 'usuario_pesagem', self.ent_usuario.get().strip())
        self.config.set(secao, 'cliente_estoque', str(self._cliente_selecionado() or ''))
        self.config.set(secao, 'uma_ordem', 'S' if self.var_uma_ordem.get() else 'N')
        self.config.set(secao, 'busca_produto', self._modo_busca_produto())
        with open('config.ini', 'w', encoding='utf-8') as f:
            self.config.write(f)

    def _carregar_config_mapeamento(self):
        secao = 'IMPORTACAO_ESTOQUE_PRODUCAO'
        if not self.config.has_section(secao):
            return
        arquivo = self.config.get(secao, 'ultimo_arquivo', fallback='')
        if arquivo and os.path.isfile(arquivo):
            self.caminho_arquivo = arquivo
            self.ent_arquivo.delete(0, tk.END)
            self.ent_arquivo.insert(0, arquivo)
            abas = obter_abas_planilha(arquivo)
            self.cb_abas['values'] = abas
            aba_salva = self.config.get(secao, 'ultima_aba', fallback='')
            if aba_salva and aba_salva in abas:
                self.cb_abas.set(aba_salva)
            elif abas:
                self.cb_abas.current(0)
        linha = self.config.get(secao, 'linha_inicial', fallback='2')
        self.ent_linha_ini.delete(0, tk.END)
        self.ent_linha_ini.insert(0, linha)
        for chave in self.entradas_map:
            valor = self.config.get(secao, f'map_{chave}', fallback='')
            if valor:
                self.entradas_map[chave].delete(0, tk.END)
                self.entradas_map[chave].insert(0, valor)
        for ent, chave, padrao in (
            (self.ent_empresa, 'empresa', None),
            (self.ent_filial, 'filial', None),
            (self.ent_obs, 'observacao', OBS_PADRAO),
            (self.ent_usuario, 'usuario_pesagem', USUARIO_PESAGEM_PADRAO),
        ):
            valor = self.config.get(secao, chave, fallback='')
            if valor:
                ent.delete(0, tk.END)
                ent.insert(0, valor)
            elif padrao and not ent.get().strip():
                ent.insert(0, padrao)
        status = self.config.get(secao, 'cod_status', fallback='')
        if status in ('P', 'A', 'F'):
            self.cb_status.set(status)
        self.var_uma_ordem.set(self.config.get(secao, 'uma_ordem', fallback='S').upper() != 'N')
        self._cliente_salvo = self.config.get(secao, 'cliente_estoque', fallback='').strip()
        modo = self.config.get(secao, 'busca_produto', fallback='auto').strip()
        for i, (lbl, chave) in enumerate(BUSCA_PRODUTO):
            if chave == modo:
                self.cb_busca_produto.current(i)
                break

    # ------------------------------------------------------------ CLIENTES
    def _carregar_clientes_bg(self):
        threading.Thread(target=self._carregar_clientes, daemon=True).start()

    def _carregar_clientes(self):
        try:
            emp = self.ent_empresa.get().strip() or '1'
            fil = self.ent_filial.get().strip() or '1'
            with FirebirdService(self.config_db) as fb:
                rows = fb.query(
                    "SELECT CF_CODIGO, CF_RAZAO FROM TABELA_CLI_FOR "
                    "WHERE CF_EMPRESA = ? AND CF_FILIAL = ? ORDER BY CF_CODIGO",
                    [emp, fil])
            clientes = [(int(r['cf_codigo']), str(r['cf_razao'] or '').strip()) for r in rows]
            self.parent.after(0, lambda c=clientes: self._popular_clientes(c))
        except Exception:
            # Sem banco disponivel a tela continua utilizavel (o codigo pode ser digitado no config)
            pass

    def _popular_clientes(self, clientes):
        self._clientes = clientes
        self.cb_cliente['values'] = [f"{cod} - {razao}" for cod, razao in clientes]
        alvo = getattr(self, '_cliente_salvo', '') or '1'
        for i, (cod, _) in enumerate(clientes):
            if str(cod) == str(alvo):
                self.cb_cliente.current(i)
                return
        if clientes:
            self.cb_cliente.current(0)

    def _modo_busca_produto(self):
        """Chave interna do modo escolhido no combo ('auto', 'codigo', ...)."""
        rotulo = self.cb_busca_produto.get()
        for lbl, chave in BUSCA_PRODUTO:
            if lbl == rotulo:
                return chave
        return 'auto'

    def _descr_modo(self, modo):
        if modo == 'auto':
            return "código do ERP, código de importação ou auxiliar"
        return {'codigo': "código do produto (ERP)",
                'importacao': "código de importação",
                'auxiliar': "código auxiliar"}.get(modo, modo)

    def _cliente_selecionado(self):
        texto = self.cb_cliente.get().strip()
        if not texto:
            return None
        try:
            return int(texto.split('-', 1)[0].strip())
        except ValueError:
            return None

    # ------------------------------------------------------------ ARQUIVO
    def _selecionar_arquivo(self):
        path = filedialog.askopenfilename(
            filetypes=[("Arquivos Suportados", "*.xlsx *.csv"), ("Excel", "*.xlsx"), ("CSV", "*.csv")])
        if path:
            self.ent_arquivo.delete(0, tk.END)
            self.ent_arquivo.insert(0, path)
            self.caminho_arquivo = path
            abas = obter_abas_planilha(path)
            self.cb_abas['values'] = abas
            if abas:
                self.cb_abas.current(0)

    def _fechar_tela(self):
        self.destroy()
        if self.callback_voltar:
            self.callback_voltar()

    # ------------------------------------------------------------ PARSERS
    def _float_br(self, valor):
        """Converte '1.234,567' / '1234.567' / '12,5' em float. Retorna None se vazio/invalido."""
        if valor is None:
            return None
        s = str(valor).strip().replace('R$', '').replace(' ', '')
        if not s:
            return None
        s = re.sub(r'[^\d,.\-]', '', s)
        if not s or s in ('-', '.', ','):
            return None
        if ',' in s and '.' in s:
            # o ultimo separador manda: 1.234,56 (BR) ou 1,234.56 (US)
            if s.rfind(',') > s.rfind('.'):
                s = s.replace('.', '').replace(',', '.')
            else:
                s = s.replace(',', '')
        elif ',' in s:
            s = s.replace(',', '.')
        try:
            return float(s)
        except ValueError:
            return None

    # tabela/coluna de cada generator, para conferir se ele esta a frente do MAX
    GENERATORS = {
        'GEN_COD_ID': ('CONFRI_ORDEM_DESOSSA', 'COD_ID'),
        'GEN_CDA_ID': ('CONFRI_ORDEM_DESOSSA_PA', 'CDA_ID'),
        'GEN_CPP_ID': ('CONFRI_ORDEM_DESOSSA_PA_PESAGEM', 'CPP_ID'),
    }

    def _sincronizar_generator(self, cur, gen):
        """Empurra o generator para frente do MAX(id) real da tabela.

        O ERP grava nessas mesmas tabelas e nem sempre pelo generator; quando o
        generator fica atrasado, GEN_ID devolve um ID já ocupado e o INSERT morre
        com "violation of PRIMARY or UNIQUE KEY" (SQLCODE -803).
        Retorna (valor_do_generator, quanto_avancou).
        """
        tabela, coluna = self.GENERATORS[gen]
        cur.execute(f"SELECT COALESCE(MAX({coluna}), 0) FROM {tabela}")
        maximo = int(cur.fetchone()[0] or 0)
        cur.execute(f"SELECT GEN_ID({gen}, 0) FROM RDB$DATABASE")
        atual = int(cur.fetchone()[0] or 0)
        if atual < maximo:
            cur.execute(f"SELECT GEN_ID({gen}, {maximo - atual}) FROM RDB$DATABASE")
            cur.fetchone()
            return maximo, maximo - atual
        return atual, 0

    def _qtde_pecas(self, valor):
        """Quantidade de peças (CPP_QTDE) — NUNCA pode ir 0 ou NULL para o banco.

        Vazio, zero, negativo ou texto inválido viram 1.
        """
        n = self._float_br(valor)
        if n is None or n <= 0:
            return 1
        return n

    def _chaves_produto(self, valor, degradar=True):
        """Formas de busca de um código de produto vindo da planilha/ERP.

        Trata célula numérica do Excel ('10000.0' -> '10000'), zeros à esquerda,
        separador de milhar e códigos alfanuméricos (COD_IMPORTACAO é VARCHAR).

        `degradar=False` deixa de fora a última forma, "só os dígitos". Ela é uma
        chave DEGRADADA e só pode ser confrontada com o valor EXATO do outro lado
        — nunca com a degradação do outro lado. Indexando o ERP com ela, códigos
        que diferem apenas nas letras passavam a dividir a mesma chave: `MCC600`
        (Maiale Canastra) e `BKC600` (Berkshire) viravam ambos '600', e o `MDC600`
        do XML (Maiale Duroc, um terceiro produto) casava com os dois —
        *AMBÍGUO (2: 207, 307)*, um empate que não existe. Pior: com só um deles
        cadastrado não haveria empate nenhum e a nota entraria calada no produto
        errado. Nessas linhas as letras são a linha do produto, não enfeite: os
        códigos terminados em `000` (SNV000, CSR000, KIT000) colidiam todos.
        """
        s = str(valor if valor is not None else '').strip().upper()
        if not s:
            return []
        formas = []

        def add(x):
            x = str(x).strip().upper()
            if not x:
                return
            x = x.lstrip('0') or '0'
            if x not in formas:
                formas.append(x)

        # 1) artefato de celula numerica do Excel: 10000.0 / 10000,00 -> 10000
        m = re.match(r'^(\d+)[.,]0{1,2}$', s)
        if m:
            add(m.group(1))
        # 2) separador de milhar: 10.000 -> 10000 / 1.234.567 -> 1234567
        elif re.match(r'^\d{1,3}([.,]\d{3})+$', s):
            add(re.sub(r'[.,]', '', s))
        # 3) o valor exatamente como veio (COD_IMPORTACAO/AUXILIAR sao VARCHAR)
        add(s)
        # 4) por ultimo, so os digitos (tolera 'AB-12' -> '12'). Fora do indice.
        if degradar:
            somente_digitos = re.sub(r'\D', '', s)
            if somente_digitos:
                add(somente_digitos)
        return formas

    def _indexar_produtos(self, rows):
        """Monta um índice por código do ERP, por código de importação e por auxiliar.

        Cada chave guarda a lista de produtos que a possuem, para detectar ambiguidade.

        O índice guarda só as formas FIÉIS do código (`degradar=False`): a forma
        "só os dígitos" existe para o lado da busca e, dos dois lados, produzia
        empate entre produtos de linhas diferentes (ver `_chaves_produto`).
        """
        indices = {'codigo': {}, 'importacao': {}, 'auxiliar': {}}
        for r in rows:
            try:
                codigo = int(r['produto_codigo'])
            except (TypeError, ValueError):
                continue
            info = {
                'codigo': codigo,
                'descricao': str(r['produto_descricao'] or '').strip(),
                'ativo': str(r['produto_ativo'] or '').strip().upper(),
                'cod_importacao': str(r.get('produto_cod_importacao') or '').strip(),
                'cod_auxiliar': str(r.get('produto_cod_auxiliar') or '').strip(),
                # opcionais: só quem seleciona estas colunas (a importação de notas)
                # os traz. Servem para conferir o produto contra o XML.
                'ncm': str(r.get('produto_class_fiscal') or '').strip(),
                'unidade': str(r.get('produto_unidade_cv')
                               or r.get('produto_unidade_est') or '').strip(),
                'cbarra': str(r.get('produto_cbarra') or '').strip(),
            }
            for campo, bruto in (('codigo', codigo),
                                 ('importacao', info['cod_importacao']),
                                 ('auxiliar', info['cod_auxiliar'])):
                for chave in self._chaves_produto(bruto, degradar=False):
                    lista = indices[campo].setdefault(chave, [])
                    if all(x['codigo'] != codigo for x in lista):
                        lista.append(info)
        return indices

    def _resolver_produto(self, valor, indices, modo):
        """Acha o produto pela coluna da planilha.

        A CHAVE vem antes do CAMPO no laço, e isso é essencial: `_chaves_produto`
        devolve as formas da mais fiel para a mais degradada, e a última delas é
        "só os dígitos". Percorrendo campo a campo primeiro, o código de importação
        'MDC131' era reduzido a '131' e casava com o PRODUTO_CODIGO 131 — outro
        produto — antes de o índice de importação ser consultado. Isso vinculou
        1.221 dos 3.540 itens das notas do FABENE ao produto errado. Um acerto
        exato em qualquer campo tem de vencer um acerto degradado em qualquer outro.

        PRODUTO INATIVO NÃO CONCORRE. Inativar o gêmeo é justamente como o usuário
        diz qual dos dois não usar — o sistema do cliente emite o mesmo código para
        dois cadastros, ele inativa o que sobrou. Contando os dois, o CORACAO 3345
        (10381 ativo, 20000 inativo) virava "AMBÍGUO (2 produtos)" e a inativação
        não servia para nada. Só quando sobra mais de um ATIVO é que há ambiguidade
        de verdade. Se todos estiverem inativos, ainda assim usa — achar o produto
        inativo é melhor que mandar cadastrar uma terceira via do mesmo item.

        Retorna (info, campo_que_casou, codigos_ambiguos).
        """
        ordem = ['codigo', 'importacao', 'auxiliar'] if modo == 'auto' else [modo]
        chaves = self._chaves_produto(valor)
        if not chaves:
            return None, None, None
        exatas = set(self._chaves_produto(valor, degradar=False))
        for chave in chaves:
            for campo in ordem:
                achados = indices.get(campo, {}).get(chave)
                if not achados:
                    continue
                ativos = [a for a in achados if a.get('ativo') != 'N']
                candidatos = ativos or achados
                codigos = sorted({a['codigo'] for a in candidatos})
                if len(codigos) > 1:
                    return None, campo, codigos
                escolhido = candidatos[0]
                if len(achados) > len(candidatos) or chave not in exatas:
                    escolhido = dict(escolhido)
                if len(achados) > len(candidatos):
                    # deixa rastro de que houve gêmeo, para o motivo da grade poder dizer
                    escolhido['gemeos_inativos'] = sorted(
                        {a['codigo'] for a in achados} - {escolhido['codigo']})
                if chave not in exatas:
                    # casou só pelos dígitos ('MDC131' -> 131): é palpite, não
                    # acerto. Quem olha a grade tem de saber para poder conferir.
                    escolhido['casou_degradado'] = chave
                return escolhido, campo, None
        return None, None, None

    def _data_br(self, valor):
        """Converte texto/serial em date. Retorna None se vazio/invalido."""
        if valor is None:
            return None
        s = str(valor).strip()
        if not s:
            return None
        # openpyxl entrega datetime que virou string no leitor
        if len(s) >= 10 and s[4] == '-' and s[7] == '-':
            try:
                return datetime.date(int(s[0:4]), int(s[5:7]), int(s[8:10]))
            except ValueError:
                return None
        m = re.match(r'^(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2,4})', s)
        if m:
            d, mo, a = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if a < 100:
                a += 2000
            try:
                return datetime.date(a, mo, d)
            except ValueError:
                return None
        # serial do Excel
        try:
            n = float(s.replace(',', '.'))
            if 1 < n < 100000:
                return datetime.date(1899, 12, 30) + datetime.timedelta(days=int(n))
        except ValueError:
            pass
        return None

    def _data_hora(self, valor, padrao):
        """Converte texto em datetime; usa `padrao` quando a coluna nao foi mapeada/esta vazia."""
        if valor is None or not str(valor).strip():
            return padrao
        s = str(valor).strip()
        for fmt in ('%Y-%m-%d %H:%M:%S', '%d/%m/%Y %H:%M:%S', '%d/%m/%Y %H:%M',
                    '%Y-%m-%d %H:%M', '%Y-%m-%dT%H:%M:%S'):
            try:
                return datetime.datetime.strptime(s[:len('2026-01-01 00:00:00')], fmt)
            except ValueError:
                continue
        d = self._data_br(s)
        if d:
            return datetime.datetime(d.year, d.month, d.day, 8, 0, 0)
        return padrao

    def _dias_validade(self, producao, validade):
        """Shelf life em dias: validade - producao. Só informativo, nao vai pro ERP.
        Retorna None quando falta uma das datas."""
        if not producao or not validade:
            return None
        return (validade - producao).days

    # ------------------------------------------------------------ ANALISE
    def _iniciar_analise(self):
        aba = self.cb_abas.get()
        try:
            linha_ini = int(self.ent_linha_ini.get())
        except ValueError:
            return messagebox.showerror("Erro", "A linha inicial deve ser um número.")

        if not self.caminho_arquivo or not aba:
            return messagebox.showwarning("Aviso", "Selecione o arquivo e a aba antes de continuar.")

        mapa_colunas = {chave: ent.get().strip() for chave, ent in self.entradas_map.items()}
        faltando = [lbl for lbl, chave, obrig in CAMPOS_MAPEAMENTO if obrig and not mapa_colunas.get(chave)]
        if faltando:
            return messagebox.showwarning(
                "Aviso", "Mapeie obrigatoriamente:\n\n- " + "\n- ".join(faltando))

        if not self.ent_empresa.get().strip().isdigit() or not self.ent_filial.get().strip().isdigit():
            return messagebox.showwarning("Aviso", "Empresa e Filial devem ser números.")

        if self._data_br(self.ent_data_ordem.get()) is None:
            return messagebox.showwarning("Aviso", "Data da Ordem inválida. Use o formato dd/mm/aaaa.")

        self._salvar_config_mapeamento()
        self.btn_analisar.config(state=tk.DISABLED)
        self.btn_importar.config(state=tk.DISABLED)
        self.btn_exportar.config(state=tk.DISABLED)
        self.btn_exportar_faltantes.config(state=tk.DISABLED)
        self.lbl_status.config(text="Lendo planilha...")
        self.progresso['value'] = 20

        threading.Thread(target=self._analisar_bg, args=(aba, mapa_colunas, linha_ini), daemon=True).start()

    def _analisar_bg(self, aba, mapa_colunas, linha_ini):
        try:
            raw = ler_planilha_produtos(self.caminho_arquivo, aba, mapa_colunas, linha_ini)
            self.registros_lidos = raw[:]

            emp = self.ent_empresa.get().strip()
            fil = self.ent_filial.get().strip()

            try:
                with FirebirdService(self.config_db) as fb:
                    rows = fb.query(
                        "SELECT PRODUTO_CODIGO, PRODUTO_DESCRICAO, PRODUTO_ATIVO, "
                        "       PRODUTO_COD_IMPORTACAO, PRODUTO_COD_AUXILIAR "
                        "FROM TABELA_PRODUTO WHERE PRODUTO_EMPRESA = ? AND PRODUTO_FILIAL = ?",
                        [emp, fil])
                    indices = self._indexar_produtos(rows)

                    rows_eti = fb.query(
                        "SELECT P.CPP_COD_ETIQUETA FROM CONFRI_ORDEM_DESOSSA_PA_PESAGEM P "
                        "INNER JOIN CONFRI_ORDEM_DESOSSA_PA A ON A.CDA_ID = P.CPP_CDA_ID "
                        "INNER JOIN CONFRI_ORDEM_DESOSSA O ON O.COD_ID = A.CDA_COD_ID "
                        "WHERE O.COD_EMPRESA = ? AND O.COD_FILIAL = ?",
                        [emp, fil])
                    etiquetas_erp = {str(r['cpp_cod_etiqueta'] or '').strip().upper()
                                     for r in rows_eti if r['cpp_cod_etiqueta']}
            except Exception as e:
                self.parent.after(0, lambda err=e: messagebox.showerror("Erro", f"Falha ao consultar ERP:\n{err}"))
                self.parent.after(0, lambda: self.btn_analisar.config(state=tk.NORMAL))
                return

            self.parent.after(0, lambda: self._renderizar_preview(indices, etiquetas_erp))
        except Exception as e:
            self.parent.after(0, lambda err=e: messagebox.showerror("Erro", f"Falha na leitura da planilha:\n{err}"))
            self.parent.after(0, lambda: self.btn_analisar.config(state=tk.NORMAL))

    def _renderizar_preview(self, indices, etiquetas_erp):
        # Selo deste render. A grade e preenchida em blocos com after(), entao um
        # render antigo pode continuar inserindo DEPOIS que outro limpou a tela —
        # a grade acumula duas analises e os totais somam tudo. O selo faz os
        # blocos do render antigo pararem.
        self._render_seq = getattr(self, '_render_seq', 0) + 1
        meu_seq = self._render_seq

        for i in self.tree.get_children():
            self.tree.delete(i)
        self.dados_grid.clear()

        agora = datetime.datetime.now()
        vistas_planilha = set()
        items = []
        validos = 0
        modo_busca = self._modo_busca_produto()

        for reg in self.registros_lidos:
            etiqueta = str(reg.get('etiqueta', '')).strip()
            prod_raw = str(reg.get('produto', '')).strip()
            peso = self._float_br(reg.get('peso'))
            peso_bruto = self._float_br(reg.get('peso_bruto'))
            qtde = self._qtde_pecas(reg.get('qtde'))
            tara_emb = self._float_br(reg.get('tara_emb'))
            tara_caixa = self._float_br(reg.get('tara_caixa'))
            validade = self._data_br(reg.get('validade'))
            producao = self._data_br(reg.get('producao'))
            data_hora = self._data_hora(reg.get('data_hora'), agora)
            descricao_pl = str(reg.get('descricao', '')).strip()

            if not etiqueta and not prod_raw and peso is None:
                continue  # linha em branco

            prod_info, campo_casou, ambiguos = self._resolver_produto(prod_raw, indices, modo_busca)

            # cadeia de validacao: o primeiro problema encontrado define o status
            if not etiqueta:
                status, obs = "ERRO", "Etiqueta em branco"
            elif len(etiqueta) > 50:
                status, obs = "ERRO", "Etiqueta acima de 50 caracteres"
            elif not prod_raw:
                status, obs = "ERRO", "Produto em branco"
            elif ambiguos:
                status, obs = ("PRODUTO AMBÍGUO",
                               f"'{prod_raw}' casa com {len(ambiguos)} produtos pelo "
                               f"{ROTULO_ORIGEM.get(campo_casou, campo_casou)}: "
                               f"{', '.join(str(c) for c in ambiguos[:6])}")
            elif prod_info is None:
                status, obs = ("PRODUTO NÃO ENCONTRADO",
                               f"'{prod_raw}' não achado por {self._descr_modo(modo_busca)} "
                               f"na empresa/filial")
            elif peso is None or peso <= 0:
                status, obs = "ERRO", "Peso inválido ou zerado"
            elif etiqueta.upper() in vistas_planilha:
                status, obs = "DUPLICADA", "Etiqueta repetida na própria planilha"
            elif etiqueta.upper() in etiquetas_erp:
                status, obs = "JÁ CADASTRADA", "Etiqueta já existe no ERP"
            else:
                status, obs = "OK", ""

            if status == "OK":
                vistas_planilha.add(etiqueta.upper())
                validos += 1

            reg['_etiqueta'] = etiqueta
            reg['_produto'] = prod_info['codigo'] if prod_info else None
            reg['_produto_raw'] = prod_raw
            reg['_casou_por'] = ROTULO_ORIGEM.get(campo_casou, '') if prod_info else ''
            reg['_descricao_erp'] = prod_info['descricao'] if prod_info else descricao_pl
            reg['_peso'] = peso
            reg['_peso_bruto'] = peso_bruto if peso_bruto is not None else peso
            reg['_qtde'] = qtde  # ja normalizado por _qtde_pecas (nunca 0/None)
            reg['_tara_emb'] = tara_emb
            reg['_tara_caixa'] = tara_caixa
            reg['_validade'] = validade
            reg['_producao'] = producao
            reg['_dias_val'] = self._dias_validade(producao, validade)
            reg['_data_hora'] = data_hora
            reg['_status'] = status
            reg['_obs'] = obs

            check = "☑" if status == "OK" else "☐"
            if status == 'OK':
                tag = 'OK'
            elif status in ('ERRO', 'PRODUTO NÃO ENCONTRADO', 'PRODUTO AMBÍGUO'):
                tag = 'ERRO'
            else:
                tag = 'AVISO'

            items.append((
                (check, status, etiqueta or '-', prod_raw or '-',
                 str(reg['_produto']) if reg['_produto'] is not None else '-',
                 reg['_casou_por'] or '-',
                 (reg['_descricao_erp'] or '-')[:60],
                 f"{peso:.3f}".replace('.', ',') if peso is not None else '-',
                 f"{reg['_peso_bruto']:.3f}".replace('.', ',') if reg['_peso_bruto'] is not None else '-',
                 f"{reg['_qtde']:g}",
                 producao.strftime('%d/%m/%Y') if producao else '-',
                 validade.strftime('%d/%m/%Y') if validade else '-',
                 str(reg['_dias_val']) if reg['_dias_val'] is not None else '-'),
                tag, reg
            ))

        total = len(items)
        if total == 0:
            self.btn_analisar.config(state=tk.NORMAL)
            self.lbl_status.config(text="Nenhum registro válido encontrado.")
            self._atualizar_cards_resumo()
            return

        self.lbl_status.config(text=f"Renderizando tabela com {total} registros...")
        self.progresso['value'] = 92
        chunk_size = 60

        def render_chunk(start_idx):
            if meu_seq != self._render_seq:
                return  # um render mais novo assumiu a grade
            end_idx = min(start_idx + chunk_size, total)
            for i in range(start_idx, end_idx):
                values, tag, reg = items[i]
                item_id = self.tree.insert("", tk.END, values=values, tags=(tag,))
                self.dados_grid[item_id] = reg

            if end_idx < total:
                self.lbl_status.config(text=f"Renderizando {end_idx}/{total}...")
                self.update_idletasks()
                self.parent.after(5, render_chunk, end_idx)
            else:
                self.btn_analisar.config(state=tk.NORMAL)
                if validos > 0:
                    self.btn_importar.config(state=tk.NORMAL)
                self.btn_exportar.config(state=tk.NORMAL)
                n_faltantes = len(self._agrupar_faltantes())
                self.btn_exportar_faltantes.config(
                    state=tk.NORMAL if n_faltantes else tk.DISABLED,
                    text=f"📦   Exportar Produtos a Cadastrar ({n_faltantes})"
                         if n_faltantes else "📦   Exportar Produtos a Cadastrar")
                self.progresso['value'] = 100
                self.lbl_status.config(text=f"Pronto. {validos} etiqueta(s) prontas de {total} lidas.")
                self.cb_filtro_status.current(0)
                self._filtrar_status()
                self._atualizar_cards_resumo()

        render_chunk(0)

    # -------------------------------------------------------------- CARDS
    def _atualizar_cards_resumo(self):
        """Soma apenas o que esta marcado para importar (o que vai de fato pro ERP)."""
        qtd = 0
        peso_total = 0.0
        produtos = set()
        problemas = 0

        for item_id, reg in self.dados_grid.items():
            marcado = False
            try:
                marcado = self.tree.item(item_id, "values")[0] == "☑"
            except tk.TclError:
                pass
            if reg.get('_status') != 'OK':
                problemas += 1
            if not marcado:
                continue
            qtd += 1
            peso_total += float(reg.get('_peso') or 0)
            if reg.get('_produto') is not None:
                produtos.add(reg['_produto'])

        self.card_etiquetas.lbl_valor.config(text=f"{qtd:,}".replace(',', '.'))
        self.card_peso.lbl_valor.config(
            text=f"{peso_total:,.3f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
        self.card_produtos.lbl_valor.config(text=str(len(produtos)))
        self.card_erros.lbl_valor.config(text=str(problemas))

    # ---------------------------------------------------------- INTERACAO
    def _on_tree_click(self, event):
        region = self.tree.identify_region(event.x, event.y)
        if region == "cell":
            column = self.tree.identify_column(event.x)
            item_id = self.tree.identify_row(event.y)
            if not item_id:
                return
            if column == "#1":
                valores = list(self.tree.item(item_id, 'values'))
                valores[0] = "☑" if valores[0] == "☐" else "☐"
                self.tree.item(item_id, values=valores)
                self._atualizar_cards_resumo()
            self._update_import_button()

    def _copiar(self, texto, descricao):
        """Coloca o texto na area de transferencia do Windows."""
        texto = str(texto or '').strip()
        if not texto or texto == '-':
            self.lbl_status.config(text="Nada para copiar nessa célula.")
            return
        try:
            self.clipboard_clear()
            self.clipboard_append(texto)
            self.update()  # sem isso o conteudo pode se perder ao fechar a janela
            resumo = texto.replace('\t', ' | ').replace('\n', ' ⏎ ')
            if len(resumo) > 60:
                resumo = resumo[:57] + '...'
            self.lbl_status.config(text=f"Copiado — {descricao}: {resumo}")
        except tk.TclError:
            pass

    def _on_tree_double_click(self, event):
        """Duplo-clique copia o valor da célula (a etiqueta, no caso da coluna ETIQUETA)."""
        if self.tree.identify_region(event.x, event.y) != "cell":
            return
        item_id = self.tree.identify_row(event.y)
        if not item_id:
            return
        coluna = self.tree.identify_column(event.x)  # '#1', '#2', ...
        try:
            idx = int(coluna.replace('#', '')) - 1
        except ValueError:
            return
        if idx <= 0 or idx >= len(self.colunas):
            return  # coluna 0 e o checkbox de seleção
        valores = self.tree.item(item_id, 'values')
        if idx >= len(valores):
            return
        self.tree.selection_set(item_id)
        self._copiar(valores[idx], self.colunas[idx])

    def _on_ctrl_c(self, event=None):
        """Ctrl+C copia a(s) linha(s) selecionada(s) inteira(s), separadas por TAB."""
        sel = self.tree.selection()
        if not sel:
            return
        linhas = []
        for item_id in sel:
            vals = list(self.tree.item(item_id, 'values'))[1:]  # sem o checkbox
            linhas.append("\t".join(str(v) for v in vals))
        self._copiar("\n".join(linhas), f"{len(sel)} linha(s)")
        return "break"

    def _update_import_button(self):
        has_checked = any(
            self.tree.item(item, "values")[0] == "☑"
            for item in self.tree.get_children())
        self.btn_importar.config(state=tk.NORMAL if has_checked else tk.DISABLED)

    def _marcar_todos(self):
        for item in self.tree.get_children():
            v = list(self.tree.item(item, "values"))
            v[0] = "☑"
            self.tree.item(item, values=v)
        self._update_import_button()
        self._atualizar_cards_resumo()

    def _desmarcar_todos(self):
        for item in self.tree.get_children():
            v = list(self.tree.item(item, "values"))
            v[0] = "☐"
            self.tree.item(item, values=v)
        self._update_import_button()
        self._atualizar_cards_resumo()

    def _sort_treeview(self, col):
        self._sort_directions[col] = not self._sort_directions[col]
        reverse = self._sort_directions[col]
        l = [(self.tree.set(k, col), k) for k in self.tree.get_children('')]

        def valor_para_ordenar(val):
            v = str(val).strip()
            if not v or v == '-':
                return -999999 if reverse else 999999
            try:
                return float(v.replace('.', '').replace(',', '.'))
            except ValueError:
                return v.lower()

        l.sort(key=lambda t: valor_para_ordenar(t[0]), reverse=reverse)
        for index, (_, k) in enumerate(l):
            self.tree.move(k, '', index)
        for c in self.colunas:
            arrow = " ▼" if self._sort_directions[c] else " ▲" if c == col else " ↕"
            self.tree.heading(c, text=c + arrow, command=lambda x=c: self._sort_treeview(x))

    def _filtrar_status(self, event=None):
        filtro = self.cb_filtro_status.get()
        for item in self.tree.get_children():
            self.tree.detach(item)
        for item_id, reg in self.dados_grid.items():
            status = reg.get('_status', '')
            if filtro == "Todos" or status == filtro:
                self.tree.move(item_id, '', tk.END)

    # ----------------------------------------------------------- EXPORTAR
    COLUNAS_EXPORT = [
        "SEQ", "MARCADO", "STATUS", "MOTIVO", "ETIQUETA", "PROD_PLANILHA",
        "COD_ERP", "CASOU_POR", "DESCRICAO", "PESO", "PESO_BRUTO", "QTDE",
        "TARA_EMB", "TARA_CAIXA", "DATA_HORA_PESAGEM", "PRODUCAO", "VENCIMENTO",
        "DIAS_VALIDADE",
    ]

    def _linhas_export(self):
        """Monta as linhas da exportação na ordem/filtro que está na tela."""
        linhas = []
        for seq, item_id in enumerate(self.tree.get_children(), start=1):
            reg = self.dados_grid.get(item_id)
            if not reg:
                continue
            marcado = self.tree.item(item_id, "values")[0] == "☑"
            linhas.append([
                seq,
                "SIM" if marcado else "NAO",
                reg.get('_status', ''),
                reg.get('_obs', ''),
                reg.get('_etiqueta', ''),
                reg.get('_produto_raw', ''),
                reg.get('_produto') if reg.get('_produto') is not None else '',
                reg.get('_casou_por', ''),
                reg.get('_descricao_erp', ''),
                reg.get('_peso'),
                reg.get('_peso_bruto'),
                reg.get('_qtde'),
                reg.get('_tara_emb'),
                reg.get('_tara_caixa'),
                reg.get('_data_hora'),
                reg.get('_producao'),
                reg.get('_validade'),
                reg.get('_dias_val'),
            ])
        return linhas

    def _resumo_export(self, linhas):
        """Contagem por status + totais, para o rodapé/aba de resumo."""
        por_status = {}
        marcadas = 0
        peso_marcado = 0.0
        produtos = set()
        for l in linhas:
            st = l[2] or '(vazio)'
            d = por_status.setdefault(st, {'qtd': 0, 'peso': 0.0})
            d['qtd'] += 1
            d['peso'] += float(l[9] or 0)
            if l[1] == "SIM":
                marcadas += 1
                peso_marcado += float(l[9] or 0)
                if l[6] != '':
                    produtos.add(l[6])
        return por_status, marcadas, peso_marcado, len(produtos)

    def _exportar_analise(self):
        linhas = self._linhas_export()
        if not linhas:
            return messagebox.showwarning("Aviso", "Não há nada analisado para exportar.")

        filtro = self.cb_filtro_status.get()
        sufixo = '' if filtro == 'Todos' else '_' + re.sub(r'\W+', '_', filtro).strip('_')
        caminho = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=f"ANALISE_ESTOQUE_PRODUCAO{sufixo}.xlsx",
            filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv")])
        if not caminho:
            return
        try:
            if caminho.lower().endswith('.csv'):
                self._exportar_csv(caminho, linhas)
            else:
                self._exportar_xlsx(caminho, linhas)
        except PermissionError:
            return messagebox.showerror(
                "Erro", "Não consegui gravar o arquivo.\n\n"
                        "Se ele estiver aberto no Excel, feche e tente de novo.", parent=self)
        except Exception as e:
            return messagebox.showerror("Erro", f"Falha ao exportar:\n{e}", parent=self)

        obs_filtro = "" if filtro == 'Todos' else f"\n\nFiltro aplicado: {filtro}"
        messagebox.showinfo("Exportado", f"{len(linhas)} linha(s) exportada(s) para:\n{caminho}{obs_filtro}",
                            parent=self)
        if messagebox.askyesno("Abrir", "Deseja abrir o arquivo agora?", parent=self):
            try:
                os.startfile(caminho)
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao abrir o arquivo:\n{e}", parent=self)

    def _num_br(self, valor, decimais=3):
        """Formata SÓ o número no padrão BR (1.234,567). Não usar replace na frase toda."""
        try:
            s = f"{float(valor):,.{decimais}f}"
        except (TypeError, ValueError):
            return str(valor)
        return s.replace(',', 'X').replace('.', ',').replace('X', '.')

    def _fmt_br(self, v):
        """Número com vírgula decimal / data dd/mm/aaaa, para o CSV abrir certo no Excel BR."""
        if v is None or v == '':
            return ''
        if isinstance(v, datetime.datetime):
            return v.strftime('%d/%m/%Y %H:%M:%S')
        if isinstance(v, datetime.date):
            return v.strftime('%d/%m/%Y')
        if isinstance(v, float):
            return f"{v:.3f}".replace('.', ',')
        return str(v)

    def _exportar_csv(self, caminho, linhas):
        with open(caminho, 'w', newline='', encoding='utf-8-sig') as f:
            w = csv.writer(f, delimiter=';')
            w.writerow(self.COLUNAS_EXPORT)
            for l in linhas:
                w.writerow([self._fmt_br(v) for v in l])
            por_status, marcadas, peso_marcado, n_prod = self._resumo_export(linhas)
            w.writerow([])
            w.writerow(['RESUMO'])
            w.writerow(['STATUS', 'QTDE', 'PESO'])
            for st in sorted(por_status):
                w.writerow([st, por_status[st]['qtd'], self._fmt_br(por_status[st]['peso'])])
            w.writerow([])
            w.writerow(['MARCADAS PARA IMPORTAR', marcadas])
            w.writerow(['PESO DAS MARCADAS', self._fmt_br(peso_marcado)])
            w.writerow(['PRODUTOS (ITENS PA)', n_prod])

    def _exportar_xlsx(self, caminho, linhas):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        cores = {
            'OK': 'EAFAF1',
            'JÁ CADASTRADA': 'FFF8E1',
            'DUPLICADA': 'FFF8E1',
            'PRODUTO NÃO ENCONTRADO': 'FDECEA',
            'PRODUTO AMBÍGUO': 'FDECEA',
            'ERRO': 'FDECEA',
        }
        wb = Workbook()
        ws = wb.active
        ws.title = "ANALISE"

        for i, nome in enumerate(self.COLUNAS_EXPORT, start=1):
            c = ws.cell(row=1, column=i, value=nome)
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = PatternFill("solid", fgColor="14146E")
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(self.COLUNAS_EXPORT))}{len(linhas) + 1}"

        for r, l in enumerate(linhas, start=2):
            fill = cores.get(l[2])
            for cc, v in enumerate(l, start=1):
                cel = ws.cell(row=r, column=cc)
                # datas e números vão como valor nativo (dá pra somar/filtrar no Excel)
                if isinstance(v, datetime.datetime):
                    cel.value = v
                    cel.number_format = 'DD/MM/YYYY HH:MM:SS'
                elif isinstance(v, datetime.date):
                    cel.value = v
                    cel.number_format = 'DD/MM/YYYY'
                elif isinstance(v, float):
                    cel.value = v
                    cel.number_format = '#,##0.000'
                else:
                    cel.value = v
                if fill:
                    cel.fill = PatternFill("solid", fgColor=fill)

        larguras = [6, 9, 24, 52, 24, 15, 10, 13, 34, 11, 12, 8, 11, 11, 20, 12, 12, 14]
        for i, larg in enumerate(larguras[:len(self.COLUNAS_EXPORT)], start=1):
            ws.column_dimensions[get_column_letter(i)].width = larg

        # --- aba de resumo ---
        por_status, marcadas, peso_marcado, n_prod = self._resumo_export(linhas)
        wr = wb.create_sheet("RESUMO")
        wr.cell(row=1, column=1, value="RESUMO DA ANÁLISE").font = Font(bold=True, size=13, color="14146E")
        wr.cell(row=2, column=1, value=f"Arquivo: {self.caminho_arquivo}")
        wr.cell(row=3, column=1, value=f"Aba: {self.cb_abas.get()}   |   "
                                       f"Empresa {self.ent_empresa.get()} / Filial {self.ent_filial.get()}")
        wr.cell(row=4, column=1, value=f"Casar produto por: {self._descr_modo(self._modo_busca_produto())}")
        wr.cell(row=5, column=1, value=f"Filtro na tela: {self.cb_filtro_status.get()}")

        linha = 7
        for i, nome in enumerate(("STATUS", "QTDE", "PESO"), start=1):
            c = wr.cell(row=linha, column=i, value=nome)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="14146E")
        for st in sorted(por_status):
            linha += 1
            wr.cell(row=linha, column=1, value=st)
            wr.cell(row=linha, column=2, value=por_status[st]['qtd'])
            cel = wr.cell(row=linha, column=3, value=por_status[st]['peso'])
            cel.number_format = '#,##0.000'
            fill = cores.get(st)
            if fill:
                for cc in (1, 2, 3):
                    wr.cell(row=linha, column=cc).fill = PatternFill("solid", fgColor=fill)

        linha += 2
        wr.cell(row=linha, column=1, value="MARCADAS PARA IMPORTAR").font = Font(bold=True)
        wr.cell(row=linha, column=2, value=marcadas)
        linha += 1
        wr.cell(row=linha, column=1, value="PESO DAS MARCADAS (kg)").font = Font(bold=True)
        cel = wr.cell(row=linha, column=2, value=peso_marcado)
        cel.number_format = '#,##0.000'
        linha += 1
        wr.cell(row=linha, column=1, value="PRODUTOS (ITENS PA)").font = Font(bold=True)
        wr.cell(row=linha, column=2, value=n_prod)

        wr.column_dimensions['A'].width = 30
        wr.column_dimensions['B'].width = 16
        wr.column_dimensions['C'].width = 16

        wb.save(caminho)

    # ------------------------------------------- PRODUTOS A CADASTRAR
    COLUNAS_FALTANTES = [
        "CODIGO_PLANILHA", "DESCRICAO_PLANILHA", "QTD_ETIQUETAS", "PESO_TOTAL",
        "QTD_PECAS", "ETIQUETAS_EXEMPLO", "CODIGO_NOVO_ERP", "UNIDADE", "NCM", "OBSERVACAO",
    ]

    def _agrupar_faltantes(self):
        """Consolida os PRODUTO NÃO ENCONTRADO: um registro por código, sem repetir.

        Varre TODAS as linhas analisadas (não só as visíveis), porque a lista é para
        pedir o cadastro — precisa estar completa mesmo com filtro na tela.
        """
        agrupado = {}
        for reg in self.dados_grid.values():
            if reg.get('_status') != 'PRODUTO NÃO ENCONTRADO':
                continue
            bruto = str(reg.get('_produto_raw', '')).strip()
            # a chave é o código normalizado, para '4664' e '04664' não virarem dois pedidos
            chaves = self._chaves_produto(bruto)
            chave = chaves[0] if chaves else bruto.upper()
            g = agrupado.setdefault(chave, {
                'codigo': bruto, 'descricoes': [], 'qtd': 0,
                'peso': 0.0, 'pecas': 0.0, 'exemplos': [],
            })
            desc = str(reg.get('_descricao_erp') or '').strip()
            if desc and desc != '-' and desc not in g['descricoes']:
                g['descricoes'].append(desc)
            g['qtd'] += 1
            g['peso'] += float(reg.get('_peso') or 0)
            g['pecas'] += float(reg.get('_qtde') or 0)
            eti = str(reg.get('_etiqueta') or '').strip()
            if eti and eti != '-' and len(g['exemplos']) < 3:
                g['exemplos'].append(eti)

        linhas = []
        for g in agrupado.values():
            linhas.append([
                g['codigo'],
                " | ".join(g['descricoes']),
                g['qtd'],
                round(g['peso'], 3),
                g['pecas'],
                ", ".join(g['exemplos']),
                "", "", "", "",   # colunas em branco para a empresa preencher
            ])
        # mais impactantes primeiro (mais etiquetas travadas), depois por código
        linhas.sort(key=lambda l: (-l[2], str(l[0])))
        return linhas

    def _exportar_faltantes(self):
        linhas = self._agrupar_faltantes()
        if not linhas:
            return messagebox.showinfo(
                "Nada a cadastrar",
                "Não há produtos com status 'PRODUTO NÃO ENCONTRADO' nesta análise.",
                parent=self)

        total_eti = sum(l[2] for l in linhas)
        total_peso = sum(l[3] for l in linhas)
        caminho = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile="PRODUTOS_A_CADASTRAR.xlsx",
            filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv")])
        if not caminho:
            return
        try:
            if caminho.lower().endswith('.csv'):
                self._faltantes_csv(caminho, linhas, total_eti, total_peso)
            else:
                self._faltantes_xlsx(caminho, linhas, total_eti, total_peso)
        except PermissionError:
            return messagebox.showerror(
                "Erro", "Não consegui gravar o arquivo.\n\n"
                        "Se ele estiver aberto no Excel, feche e tente de novo.", parent=self)
        except Exception as e:
            return messagebox.showerror("Erro", f"Falha ao exportar:\n{e}", parent=self)

        messagebox.showinfo(
            "Exportado",
            f"{len(linhas)} produto(s) distinto(s) a cadastrar.\n"
            f"Travando {total_eti} etiqueta(s) / {self._num_br(total_peso)} kg.\n\n{caminho}",
            parent=self)
        if messagebox.askyesno("Abrir", "Deseja abrir o arquivo agora?", parent=self):
            try:
                os.startfile(caminho)
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao abrir o arquivo:\n{e}", parent=self)

    def _faltantes_csv(self, caminho, linhas, total_eti, total_peso):
        with open(caminho, 'w', newline='', encoding='utf-8-sig') as f:
            w = csv.writer(f, delimiter=';')
            w.writerow(self.COLUNAS_FALTANTES)
            for l in linhas:
                w.writerow([self._fmt_br(v) for v in l])
            w.writerow([])
            w.writerow(['TOTAL', f'{len(linhas)} produtos distintos',
                        total_eti, self._fmt_br(total_peso)])

    def _faltantes_xlsx(self, caminho, linhas, total_eti, total_peso):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "PRODUTOS A CADASTRAR"

        ws.cell(row=1, column=1, value="PRODUTOS A CADASTRAR NO ERP").font = \
            Font(bold=True, size=14, color="14146E")
        ws.cell(row=2, column=1,
                value=f"Origem: {os.path.basename(self.caminho_arquivo)}   |   "
                      f"Empresa {self.ent_empresa.get()} / Filial {self.ent_filial.get()}   |   "
                      f"{len(linhas)} produtos, {total_eti} etiquetas, "
                      f"{self._num_br(total_peso)} kg travados")
        ws.cell(row=3, column=1,
                value="Estes códigos vieram na planilha de etiquetas mas não existem no cadastro "
                      "(nem por código, nem por código de importação, nem por auxiliar).")
        ws.cell(row=4, column=1,
                value="Preencher as colunas CODIGO_NOVO_ERP / UNIDADE / NCM após o cadastro.")
        for r in (2, 3, 4):
            ws.cell(row=r, column=1).font = Font(size=9, color="444651")

        cab = 6
        borda = Border(*[Side(style='thin', color='BFBFBF')] * 4)
        for i, nome in enumerate(self.COLUNAS_FALTANTES, start=1):
            c = ws.cell(row=cab, column=i, value=nome)
            c.font = Font(bold=True, color="FFFFFF", size=10)
            # as 4 ultimas sao para a empresa preencher -> destaque diferente
            c.fill = PatternFill("solid", fgColor="C8001E" if i > 6 else "14146E")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = borda

        for r, l in enumerate(linhas, start=cab + 1):
            for cc, v in enumerate(l, start=1):
                cel = ws.cell(row=r, column=cc)
                if isinstance(v, float):
                    cel.value = v
                    cel.number_format = '#,##0.000'
                elif isinstance(v, int):
                    cel.value = v
                else:
                    cel.value = v
                cel.border = borda
                if cc > 6:
                    cel.fill = PatternFill("solid", fgColor="FFF8E1")

        fim = cab + len(linhas)
        ws.cell(row=fim + 1, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=fim + 1, column=2, value=f"{len(linhas)} produtos distintos").font = Font(bold=True)
        ws.cell(row=fim + 1, column=3, value=total_eti).font = Font(bold=True)
        cel = ws.cell(row=fim + 1, column=4, value=total_peso)
        cel.font = Font(bold=True)
        cel.number_format = '#,##0.000'

        ws.freeze_panes = f"A{cab + 1}"
        ws.auto_filter.ref = f"A{cab}:{get_column_letter(len(self.COLUNAS_FALTANTES))}{fim}"
        for i, larg in enumerate([17, 46, 14, 13, 11, 40, 16, 11, 13, 26], start=1):
            ws.column_dimensions[get_column_letter(i)].width = larg

        wb.save(caminho)

    # ----------------------------------------------------------- IMPORTAR
    def _iniciar_importacao(self):
        selecionados = []
        for item_id in self.tree.get_children():
            if self.tree.item(item_id, "values")[0] == "☑":
                selecionados.append(self.dados_grid[item_id])

        if not selecionados:
            return messagebox.showwarning("Aviso", "Selecione pelo menos uma etiqueta para importar.")

        bloqueados = [r for r in selecionados if r.get('_status') != 'OK']
        if bloqueados:
            return messagebox.showwarning(
                "Aviso",
                f"{len(bloqueados)} etiqueta(s) marcada(s) estão com problema "
                f"(ex.: {bloqueados[0].get('_status')}).\n\n"
                "Desmarque-as ou corrija a planilha antes de importar.")

        cliente = self._cliente_selecionado()
        if cliente is None:
            return messagebox.showwarning("Aviso", "Selecione o cliente padrão de estoque.")

        produtos = {r['_produto'] for r in selecionados}
        peso_total = sum(float(r.get('_peso') or 0) for r in selecionados)

        resp = messagebox.askyesno(
            "Confirmar Importação",
            f"Serão gravados no ERP:\n\n"
            f"• {'1 ordem' if self.var_uma_ordem.get() else f'{len(produtos)} ordens'} de inventário "
            f"(CONFRI_ORDEM_DESOSSA)\n"
            f"• {len(produtos)} item(ns) PA (CONFRI_ORDEM_DESOSSA_PA)\n"
            f"• {len(selecionados)} etiqueta(s) / pesagem(ns)\n"
            f"• Peso total: {self._num_br(peso_total)} kg\n\n"
            f"Empresa {self.ent_empresa.get()} / Filial {self.ent_filial.get()} — "
            f"Cliente {cliente}\n\nConfirma?")
        if not resp:
            return

        self._salvar_config_mapeamento()
        self.btn_importar.config(state=tk.DISABLED)
        self.btn_analisar.config(state=tk.DISABLED)
        self.lbl_status.config(text="Importando estoque de produção...")
        self.progresso['value'] = 0
        threading.Thread(target=self._importacao_bg, args=(selecionados, cliente), daemon=True).start()

    def _importacao_bg(self, selecionados, cliente):
        log_linhas = []
        try:
            emp = int(self.ent_empresa.get().strip())
            fil = int(self.ent_filial.get().strip())
            data_ordem = self._data_br(self.ent_data_ordem.get())
            cod_status = self.cb_status.get().strip() or COD_STATUS_PADRAO
            obs = self.ent_obs.get().strip() or OBS_PADRAO
            usuario = self.ent_usuario.get().strip() or USUARIO_PESAGEM_PADRAO
            uma_ordem = self.var_uma_ordem.get()

            # agrupa as etiquetas por produto (um item PA por produto)
            por_produto = {}
            for reg in selecionados:
                por_produto.setdefault(reg['_produto'], []).append(reg)

            ordens = 0
            itens_pa = 0
            pesagens = 0
            erros = 0
            total = len(selecionados)
            gravadas = 0

            with FirebirdService(self.config_db) as fb:
                cur = fb.conn.cursor()

                # Antes de gravar: garante que os 3 generators estao a frente do
                # MAX(id) de cada tabela. Sem isso, se o ERP tiver gravado por fora,
                # o primeiro INSERT ja bate em PK violation.
                for _g in self.GENERATORS:
                    _valor, _avancou = self._sincronizar_generator(cur, _g)
                    if _avancou:
                        log_linhas.append(
                            f"⚠ {_g} estava atrasado — avancado em {_avancou} "
                            f"para {_valor} (o ERP gravou fora do generator)")

                def gen(nome):
                    cur.execute(f"SELECT GEN_ID({nome}, 1) FROM RDB$DATABASE")
                    return int(cur.fetchone()[0])

                def eh_pk_violation(erro):
                    t = str(erro).lower()
                    return '-803' in t or 'primary or unique key' in t

                def executar(sql, params, gen_id, idx_id=0):
                    """Executa o INSERT; se bater PK violation, ressincroniza o
                    generator, troca o ID e tenta uma vez mais."""
                    try:
                        cur.execute(sql, params)
                        return params[idx_id]
                    except Exception as e:
                        if not eh_pk_violation(e):
                            raise
                        self._sincronizar_generator(cur, gen_id)
                        novo = gen(gen_id)
                        params = list(params)
                        params[idx_id] = novo
                        cur.execute(sql, params)
                        log_linhas.append(f"⚠ ID em uso — regravado com {gen_id}={novo}")
                        return novo

                def proximo_op():
                    cur.execute(
                        "SELECT COALESCE(MAX(COD_OP), 0) + 1 FROM CONFRI_ORDEM_DESOSSA "
                        "WHERE COD_EMPRESA = ? AND COD_FILIAL = ?", [emp, fil])
                    return int(cur.fetchone()[0])

                def criar_ordem():
                    cod_id = gen('GEN_COD_ID')
                    cod_op = proximo_op()
                    cod_id = executar("""
                        INSERT INTO CONFRI_ORDEM_DESOSSA
                          (COD_ID, COD_OP, COD_DATA, COD_EMPRESA, COD_FILIAL, COD_OBS,
                           COD_STATUS, COD_TIPO, COD_CF_EMPRESA, COD_CF_FILIAL, COD_CF_CODIGO)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, [cod_id, cod_op, data_ordem, emp, fil, obs,
                          cod_status, COD_TIPO_DESOSSA, emp, fil, cliente], 'GEN_COD_ID')
                    log_linhas.append(f"✅ Ordem criada: COD_ID={cod_id} COD_OP={cod_op} "
                                      f"status={cod_status} tipo={COD_TIPO_DESOSSA} cliente={cliente}")
                    return cod_id

                cod_id_unico = criar_ordem() if uma_ordem else None
                if cod_id_unico:
                    ordens += 1
                    fb.conn.commit()
                    cur = fb.conn.cursor()

                produtos_falhos = []
                for produto, regs in por_produto.items():
                    # SAVEPOINT por produto: se este item PA falhar, desfaz SO ele.
                    # (Antes era fb.conn.rollback(), que descartava o lote inteiro —
                    #  até 300 etiquetas já gravadas de outros produtos.)
                    tem_sp = False
                    try:
                        cur.execute("SAVEPOINT SP_PA")
                        tem_sp = True
                    except Exception:
                        pass
                    try:
                        if uma_ordem:
                            cod_id = cod_id_unico
                        else:
                            cod_id = criar_ordem()
                            ordens += 1

                        cda_id = gen('GEN_CDA_ID')
                        cda_id = executar("""
                            INSERT INTO CONFRI_ORDEM_DESOSSA_PA
                              (CDA_ID, CDA_COD_ID, CDA_PROD_EMPRESA, CDA_PROD_FILIAL, CDA_PROD, CDA_QTDE)
                            VALUES (?, ?, ?, ?, ?, ?)
                        """, [cda_id, cod_id, emp, fil, produto, CDA_QTDE_PADRAO], 'GEN_CDA_ID')
                        itens_pa += 1
                        log_linhas.append(f"✅ Item PA criado: CDA_ID={cda_id} produto={produto} "
                                          f"({len(regs)} etiqueta(s))")
                    except Exception as e:
                        erros += len(regs)
                        produtos_falhos.append((produto, len(regs)))
                        log_linhas.append(f"❌ Erro ao criar item PA do produto {produto}: {e}")
                        log_linhas.append(f"   -> {len(regs)} etiqueta(s) deste produto NAO foram gravadas")
                        if tem_sp:
                            try:
                                cur.execute("ROLLBACK TO SAVEPOINT SP_PA")
                            except Exception:
                                pass
                        continue

                    for reg in regs:
                        try:
                            cpp_id = gen('GEN_CPP_ID')
                            # rede de segurança: CPP_QTDE nunca vai 0 nem NULL
                            qtde_gravar = self._qtde_pecas(reg.get('_qtde'))
                            cpp_id = executar("""
                                INSERT INTO CONFRI_ORDEM_DESOSSA_PA_PESAGEM
                                  (CPP_ID, CPP_CDA_ID, CPP_COD_ETIQUETA, CPP_PESO, CPP_DATA_HORA,
                                   CPP_PESO_BRUTO, CPP_QTDE, CPP_TARA_EMB, CPP_TARA_CAIXA,
                                   CPP_VALIDADE, CPP_PRODUCAO, CPP_USUARIO_PESAGEM)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """, [cpp_id, cda_id, reg['_etiqueta'], reg['_peso'], reg['_data_hora'],
                                  reg['_peso_bruto'], qtde_gravar, reg['_tara_emb'], reg['_tara_caixa'],
                                  reg['_validade'], reg['_producao'], usuario], 'GEN_CPP_ID')
                            pesagens += 1
                            origem = f" (planilha '{reg.get('_produto_raw')}' via {reg.get('_casou_por')})" \
                                if reg.get('_casou_por') and str(reg.get('_produto_raw')) != str(produto) else ""
                            log_linhas.append(
                                f"✅ Etiqueta {reg['_etiqueta']} | produto {produto}{origem} | "
                                f"peso {reg['_peso']} | CPP_ID={cpp_id}")
                        except Exception as e:
                            erros += 1
                            log_linhas.append(f"❌ Erro na etiqueta {reg.get('_etiqueta')}: {e}")

                        gravadas += 1
                        if gravadas % LOTE_COMMIT == 0:
                            fb.conn.commit()
                            cur = fb.conn.cursor()
                            pct = int(gravadas / total * 100) if total else 100
                            self.parent.after(0, lambda p=pct, g=gravadas, t=total: self._progresso(p, g, t))

                fb.conn.commit()  # grava o restante

            msg = (f"Importação concluída!\n\n"
                   f"{ordens} ordem(ns) de inventário criada(s).\n"
                   f"{itens_pa} item(ns) PA criado(s).\n"
                   f"{pesagens} etiqueta(s) gravada(s).")
            if produtos_falhos:
                eti_perdidas = sum(n for _, n in produtos_falhos)
                msg += (f"\n\n⚠ {len(produtos_falhos)} produto(s) falharam e "
                        f"{eti_perdidas} etiqueta(s) NÃO entraram.\n"
                        "Pode reanalisar e importar de novo: o que já está no ERP "
                        "aparece como JÁ CADASTRADA e só o que faltou entra.")
            elif erros:
                msg += f"\n{erros} erro(s). Veja o log."

            self.parent.after(0, lambda m=msg: self._safe_showinfo("Concluído", m))
            log_str = "\n".join(log_linhas)
            self.parent.after(0, lambda l=log_str: self._oferecer_log(l))
            self.parent.after(0, self._limpar_e_reiniciar)

        except Exception as e:
            self.parent.after(0, lambda err=e: self._safe_showerror(
                "Erro de Importação", f"Ocorreu um erro estrutural:\n{err}"))
        finally:
            self.parent.after(0, self._resetar_ui)

    def _progresso(self, pct, gravadas, total):
        try:
            self.progresso['value'] = pct
            self.lbl_status.config(text=f"Gravando {gravadas}/{total} etiquetas...")
        except tk.TclError:
            pass

    def _limpar_e_reiniciar(self):
        self.tree.delete(*self.tree.get_children())
        self.dados_grid.clear()
        self.btn_importar.config(state=tk.DISABLED)
        self.btn_exportar.config(state=tk.DISABLED)
        self.btn_exportar_faltantes.config(state=tk.DISABLED)
        self.lbl_status.config(text="Pronto. Aguardando nova análise...")
        self.progresso['value'] = 0
        self._atualizar_cards_resumo()

    def _safe_showinfo(self, titulo, msg):
        try:
            if self.winfo_exists():
                messagebox.showinfo(titulo, msg, parent=self)
        except tk.TclError:
            pass

    def _safe_showerror(self, titulo, msg):
        try:
            if self.winfo_exists():
                messagebox.showerror(titulo, msg, parent=self)
        except tk.TclError:
            pass

    def _resetar_ui(self):
        """Reabilita o que faz sentido: com a grade vazia, importar/exportar ficam off."""
        try:
            tem_linhas = bool(self.tree.get_children())
        except tk.TclError:
            return
        estados = ((self.btn_analisar, True),
                   (self.btn_importar, tem_linhas),
                   (self.btn_exportar, tem_linhas),
                   (self.btn_exportar_faltantes, tem_linhas and bool(self._agrupar_faltantes())))
        for btn, ligado in estados:
            try:
                if btn.winfo_exists():
                    btn.config(state=tk.NORMAL if ligado else tk.DISABLED)
            except tk.TclError:
                pass

    def _oferecer_log(self, log_str):
        if not log_str.strip():
            return
        resp = messagebox.askyesno("Log da Importação",
                                   "Deseja salvar um arquivo .txt com o log detalhado?")
        if resp:
            caminho = filedialog.asksaveasfilename(
                defaultextension=".txt",
                initialfile="LOG_IMPORTACAO_ESTOQUE_PRODUCAO.txt",
                filetypes=[("Arquivos de Texto", "*.txt")])
            if caminho:
                try:
                    with open(caminho, 'w', encoding='utf-8') as f:
                        f.write("--- LOG DE IMPORTACAO DE ESTOQUE DE PRODUCAO (PA POR ETIQUETAS) ---\n\n")
                        f.write(log_str)
                    messagebox.showinfo("Log Salvo", f"Arquivo salvo em:\n{caminho}")
                    if messagebox.askyesno("Abrir Log", "Deseja abrir o arquivo de log agora?"):
                        try:
                            os.startfile(caminho)
                        except Exception as e:
                            messagebox.showerror("Erro", f"Erro ao abrir arquivo:\n{e}")
                except Exception as e:
                    messagebox.showerror("Erro", f"Erro ao salvar log:\n{e}")
